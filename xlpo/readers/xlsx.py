# -*- coding: utf-8 -*-
from __future__ import absolute_import, unicode_literals

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from xlpo.readers import TranslationsReader
from xlpo import Translation
from zipfile import BadZipfile
import os
import six


class XLSXTranslationsReader(TranslationsReader):
    """
    Reads translations from an XLSX file.
    """

    def __init__(self, filename, sheet=0, messages_col=0, translations_col=1):
        self.filename = filename

        if not filename:
            raise Exception("No file specified")

        if not isinstance(sheet, int) or sheet < 0:
            raise Exception("Invalid sheet index: {0}".format(sheet))
        self.sheet = sheet

        if not isinstance(messages_col, int) or messages_col < 0:
            raise Exception("Invalid messages column index: {0}".format(
                messages_col))
        self.messages_col = messages_col

        if not isinstance(translations_col, int) or translations_col < 0:
            raise Exception("Invalid translations column index: {0}".format(
                translations_col))
        self.translations_col = translations_col

        super(XLSXTranslationsReader, self).__init__()

    def read(self):
        if self._translations is not None:
            return

        if not os.path.exists(self.filename):
            raise IOError("File not found: {0}".format(self.filename))

        if not os.path.isfile(self.filename):
            raise IOError("Invalid file path: {0}".format(self.filename))

        try:
            workbook = load_workbook(self.filename, read_only=True)
        except (BadZipfile, InvalidFileException):
            raise IOError(
                "Invalid or corrupted file: {0}".format(self.filename))

        if self.sheet > len(workbook.worksheets) - 1:
            raise IOError("Sheet {0:d} not found".format(self.sheet))
        worksheet = workbook.worksheets[self.sheet]

        if self.messages_col > len(worksheet.columns) - 1:
            raise IOError("Messages column {0:d} not found".format(
                self.messages_col))
        if self.translations_col > len(worksheet.columns) - 1:
            raise IOError("Translations column {0:d} not found".format(
                self.translations_col))

        self._translations = []
        seen = set()

        r = 0
        for row in worksheet.iter_rows():
            message = six.text_type(row[self.messages_col].value)
            if message in seen:
                raise Exception("Duplicate message found " +
                                "at row {0:d}".format(r))
            seen.add(message)

            translation = six.text_type(row[self.translations_col].value)

            self._translations.append(Translation(message,
                                                  translation=translation))
            r += 1

        return self._translations
